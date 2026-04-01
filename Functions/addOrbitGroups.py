#!/usr/bin/env python3

from __future__ import annotations

import argparse
import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional

import openpyxl


_ORBIT_GROUP_DIGIT_RE = re.compile(r"\d+")
_LEADING_QUANTIFIERS_RE = re.compile(r"^(forall|exists)\s+[^.]+\.\s*", re.IGNORECASE)

# Rough token extraction for the kinds of formulas we emit in this repo.
_PRED_CALL_RE = re.compile(r"~?[A-Za-z_][A-Za-z0-9_]*\([^()]*\)")
_EQ_RE = re.compile(r"\b[A-Za-z_][A-Za-z0-9_]*\d*\s*(?:~=|=)\s*[A-Za-z_][A-Za-z0-9_]*\d*\b")


def _strip_digits(text: str) -> str:
	return _ORBIT_GROUP_DIGIT_RE.sub("", text)


def _extract_literals(formula: str) -> list[str]:
	"""Extract a multiset of atom-like literals from a quantified formula string.

	The spreadsheet column contains strings like:
	  - "forall NODE0. ~locked_epoch0(NODE0)"
	  - "forall NODE0. ~ep_epoch0(NODE0) | ~held(NODE0)"
	  - "... | NODE0 = NODE1" / "... | NODE0 ~= NODE2"

	For orbit grouping we only need *which atoms appear*, not full boolean structure.
	So we pull out predicate calls and (dis)equalities as literals.
	"""

	body = _LEADING_QUANTIFIERS_RE.sub("", str(formula).strip())

	lits: list[str] = []
	lits.extend(m.group(0).strip() for m in _PRED_CALL_RE.finditer(body))
	lits.extend(" ".join(m.group(0).split()) for m in _EQ_RE.finditer(body))

	# Deterministic order for stable keys before we feed them into Counter
	lits.sort()
	return lits


def orbit_group_key_from_literals(literals: list[str]) -> Any:
	"""Prime-style orbit-group key: strip digits, keep literal signs, treat as multiset.

	This mirrors `prime.py`'s `get_orbit_group_key` fallback (string-based) logic.
	"""

	terms: list[tuple[str, str]] = []
	for lit in literals:
		lit = str(lit).strip()
		if not lit:
			continue
		if lit.startswith("~"):
			sign = "-"
			atom = lit[1:]
		else:
			sign = "+"
			atom = lit
		terms.append((sign, _strip_digits(atom)))
	return frozenset(Counter(terms).items())


@dataclass(frozen=True)
class SheetConfig:
	sheet_name: str
	formula_col: int
	header_row: int = 1


def _find_formula_column(headers: list[Any]) -> Optional[int]:
	# Prefer canonical column names observed in this repo.
	candidates = [
		"a",
		"sqi",
		"formula",
		"qclause",
		"clause",
	]
	lowered = [str(h).strip().lower() if h is not None else "" for h in headers]
	for name in candidates:
		if name in lowered:
			return lowered.index(name) + 1  # 1-based openpyxl column index
	return None


def _get_sheet_config(wb: openpyxl.Workbook, sheet: Optional[str]) -> SheetConfig:
	sheet_name = sheet or wb.sheetnames[0]
	if sheet_name not in wb.sheetnames:
		raise SystemExit(f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
	ws = wb[sheet_name]
	headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
	formula_col = _find_formula_column(headers)
	if formula_col is None:
		raise SystemExit(
			"Could not find a formula column. Expected one of: a, SQI, Formula, qclause, Clause. "
			f"Headers were: {headers}"
		)
	return SheetConfig(sheet_name=sheet_name, formula_col=formula_col)


def _ensure_output_column(ws: openpyxl.worksheet.worksheet.Worksheet, header: str) -> int:
	headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
	for idx, h in enumerate(headers, start=1):
		if str(h).strip().lower() == header.strip().lower():
			return idx
	out_col = len(headers) + 1
	ws.cell(row=1, column=out_col, value=header)
	return out_col


def add_orbit_groups(
	input_path: Path,
	*,
	output_path: Path,
	sheet: Optional[str] = None,
	output_header: str = "Orbit Group",
	size_header: str = "Orbit Group Size",
) -> None:
	wb = openpyxl.load_workbook(input_path)
	config = _get_sheet_config(wb, sheet)
	ws = wb[config.sheet_name]

	out_col = _ensure_output_column(ws, output_header)
	size_col = _ensure_output_column(ws, size_header)

	# Pass 1: compute all keys so we can assign group IDs in a deterministic order.
	keys_by_row: dict[int, Any] = {}
	key_counts: Counter[Any] = Counter()
	for row_idx in range(2, ws.max_row + 1):
		cell_val = ws.cell(row=row_idx, column=config.formula_col).value
		if cell_val is None:
			continue
		literals = _extract_literals(str(cell_val))
		key = orbit_group_key_from_literals(literals)
		keys_by_row[row_idx] = key
		key_counts[key] += 1

	# Deterministic group IDs: larger groups first, then key string.
	ordered_keys = sorted(key_counts.items(), key=lambda kv: (-kv[1], str(kv[0])))
	key_to_group_id = {k: i for i, (k, _) in enumerate(ordered_keys)}
	key_to_group_size = {k: size for (k, size) in ordered_keys}

	# Pass 2: write group id + group size per row.
	for row_idx, key in keys_by_row.items():
		ws.cell(row=row_idx, column=out_col, value=key_to_group_id[key])
		ws.cell(row=row_idx, column=size_col, value=key_to_group_size[key])

	output_path.parent.mkdir(parents=True, exist_ok=True)
	wb.save(output_path)


def main(argv: Optional[Iterable[str]] = None) -> int:
	parser = argparse.ArgumentParser(description="Add an orbit-group column to a result XLSX")
	parser.add_argument(
		"xlsx",
		type=Path,
		help="Input .xlsx (e.g. Functions/Distributed_epoch5_filtered.xlsx)",
	)
	parser.add_argument(
		"--sheet",
		type=str,
		default=None,
		help="Worksheet name to process (defaults to the first sheet)",
	)
	parser.add_argument(
		"--out",
		type=Path,
		default=None,
		help="Output .xlsx (defaults to '<input>_with_orbit_groups.xlsx')",
	)
	parser.add_argument(
		"--header",
		type=str,
		default="Orbit Group",
		help="Output column header name (default: 'Orbit Group')",
	)
	parser.add_argument(
		"--size-header",
		type=str,
		default="Orbit Group Size",
		help="Output column header name for group sizes (default: 'Orbit Group Size')",
	)
	args = parser.parse_args(list(argv) if argv is not None else None)

	input_path: Path = args.xlsx
	if not input_path.exists():
		raise SystemExit(f"Input file not found: {input_path}")

	output_path: Path
	if args.out is not None:
		output_path = args.out
	else:
		output_path = input_path.with_name(f"{input_path.stem}_with_orbit_groups{input_path.suffix}")

	add_orbit_groups(
		input_path,
		output_path=output_path,
		sheet=args.sheet,
		output_header=args.header,
		size_header=args.size_header,
	)
	print(f"Wrote: {output_path}")
	return 0


if __name__ == "__main__":
	raise SystemExit(main())

